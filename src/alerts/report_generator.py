"""Automated daily briefing report generator."""

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class BriefingGenerator:
    """Generates daily Excel briefing reports for portfolio managers."""

    def __init__(self, output_dir: str = "outputs"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(self, summaries: list[dict[str, Any]]) -> str:
        """Generate daily briefing Excel report."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Briefing"

        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = f"Daily Financial Briefing — {datetime.now().strftime('%B %d, %Y')}"
        ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")

        # Headers
        headers = ["Source", "Title", "Summary", "Sentiment", "Key Entities", "Material Events"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Data
        for row, s in enumerate(summaries[:50], 4):
            ws.cell(row=row, column=1, value=s.get("source", "")).border = border
            ws.cell(row=row, column=2, value=s.get("title", "")[:80]).border = border
            ws.cell(row=row, column=3, value=s.get("summary", "")[:200]).border = border

            sentiment = s.get("sentiment", "neutral")
            cell = ws.cell(row=row, column=4, value=sentiment)
            cell.border = border
            cell.font = Font(
                bold=True,
                color="228B22" if sentiment == "bullish" else "DC143C" if sentiment == "bearish" else "808080",
            )

            ws.cell(row=row, column=5, value=", ".join(s.get("key_entities", []))).border = border
            ws.cell(row=row, column=6, value=", ".join(s.get("material_events", []))).border = border

        # Column widths
        widths = [15, 40, 60, 12, 25, 30]
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i)].width = w

        filepath = self.output_dir / f"briefing_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filepath)
        return str(filepath)
